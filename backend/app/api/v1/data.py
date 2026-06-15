import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_role
from app.models.order import Order
from app.models.product import Product
from app.models.user import User
from app.schemas.product import ProductCreate, ProductResponse
from app.services.backup import export_backup, import_backup, export_to_storage, restore_from_storage

router = APIRouter()

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)


def _style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


@router.get("/export/products")
async def export_products(
    current_user: User = Depends(require_role("admin")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Product).where(Product.owner_id == current_user.id, Product.is_active == True))
    products = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ["ID", "Name", "SKU", "Category", "Brand", "Cost Price", "Selling Price", "Stock Quantity", "Reorder Threshold", "Batch Number", "MFG Date", "Expiry Date"]
    _style_header(ws, headers)

    for i, p in enumerate(products, 2):
        ws.cell(row=i, column=1, value=p.id)
        ws.cell(row=i, column=2, value=p.name)
        ws.cell(row=i, column=3, value=p.sku)
        ws.cell(row=i, column=4, value=p.category)
        ws.cell(row=i, column=5, value=p.brand)
        ws.cell(row=i, column=6, value=p.cost_price)
        ws.cell(row=i, column=7, value=p.selling_price)
        ws.cell(row=i, column=8, value=p.stock_quantity)
        ws.cell(row=i, column=9, value=p.reorder_threshold)
        ws.cell(row=i, column=10, value=p.batch_number)
        ws.cell(row=i, column=11, value=str(p.mfg_date) if p.mfg_date else None)
        ws.cell(row=i, column=12, value=str(p.expiry_date) if p.expiry_date else None)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=products_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/export/orders")
async def export_orders(
    current_user: User = Depends(require_role("admin")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Order).where(Order.shopkeeper_id == current_user.id).options(selectinload(Order.items)).order_by(Order.created_at.desc())
    )
    orders = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["Order #", "Status", "Payment Method", "Subtotal", "Discount", "GST", "Total", "Items", "Created At"]
    _style_header(ws, headers)

    for i, o in enumerate(orders, 2):
        items_str = "; ".join(f"{item.product_name} x{item.quantity}" for item in o.items)
        ws.cell(row=i, column=1, value=o.order_number)
        ws.cell(row=i, column=2, value=o.status.value if hasattr(o.status, 'value') else o.status)
        ws.cell(row=i, column=3, value=o.payment_method.value if hasattr(o.payment_method, 'value') else o.payment_method)
        ws.cell(row=i, column=4, value=o.subtotal)
        ws.cell(row=i, column=5, value=o.discount)
        ws.cell(row=i, column=6, value=o.gst)
        ws.cell(row=i, column=7, value=o.total)
        ws.cell(row=i, column=8, value=items_str)
        ws.cell(row=i, column=9, value=str(o.created_at))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=orders_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/backup/export")
async def backup_export(current_user: User = Depends(require_role("admin"))):
    backup = await export_backup()
    return backup


@router.post("/backup/import")
async def backup_import(payload: dict, current_user: User = Depends(require_role("admin"))):
    result = await import_backup(payload)
    return result


@router.post("/backup/export-r2")
async def backup_export_r2(current_user: User = Depends(require_role("admin"))):
    key = await export_to_storage()
    if not key:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="R2 storage not configured")
    return {"key": key}


@router.post("/backup/restore-r2")
async def backup_restore_r2(key: str, current_user: User = Depends(require_role("admin"))):
    result = await restore_from_storage(key)
    return result


@router.post("/import/products", response_model=dict)
async def import_products(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("admin")),
    db: AsyncSession = Depends(get_db),
):
    if file.filename is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No file provided")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx, .xls, and .csv files are supported")

    contents = await file.read()

    if ext == "csv":
        rows = _parse_csv(contents)
    else:
        rows = _parse_excel(contents)

    created = 0
    errors = []
    for row in rows:
        try:
            existing = await db.execute(
                select(Product).where(Product.sku == row["sku"], Product.owner_id == current_user.id)
            )
            if existing.scalar_one_or_none():
                errors.append(f"SKU '{row['sku']}' already exists")
                continue

            mfg = None
            exp = None
            if row.get("mfg_date"):
                try:
                    mfg = date.fromisoformat(str(row["mfg_date"]))
                except (ValueError, TypeError):
                    pass
            if row.get("expiry_date"):
                try:
                    exp = date.fromisoformat(str(row["expiry_date"]))
                except (ValueError, TypeError):
                    pass

            product = Product(
                name=row["name"],
                sku=row["sku"],
                category=row.get("category", "general"),
                brand=row.get("brand"),
                cost_price=float(row.get("cost_price", 0)),
                selling_price=float(row.get("selling_price", 0)),
                stock_quantity=int(row.get("stock_quantity", 0)),
                reorder_threshold=int(row.get("reorder_threshold", 10)),
                batch_number=row.get("batch_number"),
                mfg_date=mfg,
                expiry_date=exp,
                owner_id=current_user.id,
            )
            db.add(product)
            await db.flush()
            created += 1
        except Exception as e:
            errors.append(f"Row '{row.get('sku', 'unknown')}': {str(e)}")

    await db.commit()
    return {"created": created, "errors": errors, "total": len(rows)}


def _parse_excel(contents: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, [str(v) if v is not None else None for v in row])))
    return rows


def _parse_csv(contents: bytes) -> list[dict]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{k.strip(): (v.strip() if v else None) for k, v in row.items()} for row in reader]
